import datetime
import re
from typing import List
from openpyxl import load_workbook

import config
from models import LineItem

# Internal canonical name -> normalized forms it will accept in the actual
# header row (case/space/slash/underscore insensitive). Lets the sheet use
# spec-style headers like "Order ID" or "Product / SKU" instead of requiring
# the exact compact names.
COLUMNS = [
    "Platform", "OrderID", "SKU", "ReturnWindow",
    "ReturnID", "ReturnStatus", "RefundAmount",
    "TaskStatus", "Timestamp", "ErrorNote",
]

ALIASES = {
    "Platform": ["platform"],
    "OrderID": ["orderid", "order id"],
    "SKU": ["sku", "product sku", "product / sku", "product"],
    "ReturnWindow": ["returnwindow", "return window"],
    "ReturnID": ["returnid", "return id"],
    "ReturnStatus": ["returnstatus", "return status"],
    "RefundAmount": ["refundamount", "refund amount"],
    "TaskStatus": ["taskstatus", "task status"],
    "Timestamp": ["timestamp", "timestamp / log", "timestamp log",
                  "timestamp / logs", "timestamp logs"],
    "ErrorNote": ["errornote", "error note", "log", "notes"],
}


def _normalize(s: str) -> str:
    s = s.lower().strip()
    s = re.sub(r"[/_]", " ", s)
    s = re.sub(r"\s+", " ", s)
    return s


OPTIONAL_COLUMNS = {"ErrorNote"}


def _build_column_map(header: List[str]) -> dict:
    normalized_header = [_normalize(h) if h else "" for h in header]
    col = {}
    for canonical in COLUMNS:
        accepted = ALIASES[canonical]
        match_idx = None
        for i, h in enumerate(normalized_header):
            if h in accepted:
                match_idx = i
                break
        if match_idx is None:
            if canonical in OPTIONAL_COLUMNS:
                col[canonical] = None
                continue
            raise ValueError(
                f"Couldn't find a column for '{canonical}' in the header row. "
                f"Header found: {header}. Accepted names for this column: {accepted}"
            )
        col[canonical] = match_idx + 1
    return col


def read_pending_tasks() -> List[LineItem]:
    wb = load_workbook(config.EXCEL_PATH)
    ws = wb[config.SHEET_NAME]

    header = [c.value for c in ws[1]]
    col = _build_column_map(header)

    tasks = []
    for row_idx in range(2, ws.max_row + 1):
        status = ws.cell(row=row_idx, column=col["TaskStatus"]).value
        if status not in (config.STATUS_TODO, config.STATUS_PENDING, None, ""):
            continue

        tasks.append(LineItem(
            row_index=row_idx,
            platform=ws.cell(row=row_idx, column=col["Platform"]).value,
            order_id=ws.cell(row=row_idx, column=col["OrderID"]).value,
            sku=ws.cell(row=row_idx, column=col["SKU"]).value,
            return_window=ws.cell(row=row_idx, column=col["ReturnWindow"]).value,
        ))
    return tasks


def write_result(item: LineItem) -> None:
    """Write one line item's result back immediately — not batched at run end,
    so a mid-run crash doesn't lose already-completed items."""
    wb = load_workbook(config.EXCEL_PATH)
    ws = wb[config.SHEET_NAME]
    header = [c.value for c in ws[1]]
    col = _build_column_map(header)

    ws.cell(row=item.row_index, column=col["ReturnID"], value=item.return_id)
    ws.cell(row=item.row_index, column=col["ReturnStatus"], value=item.return_status)
    ws.cell(row=item.row_index, column=col["RefundAmount"], value=item.refund_amount)
    ws.cell(row=item.row_index, column=col["TaskStatus"], value=item.task_status)
    ws.cell(row=item.row_index, column=col["Timestamp"],
            value=datetime.datetime.now().isoformat(timespec="seconds"))
    if col["ErrorNote"] is not None:
        ws.cell(row=item.row_index, column=col["ErrorNote"], value=item.error_note)

    wb.save(config.EXCEL_PATH)
